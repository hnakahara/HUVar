"""Build ``resources/shared/gene_inheritance.tsv`` — a gene -> inheritance map used by
PM2 for inheritance-aware allele-frequency thresholds.

Sources, merged in priority order (later sources never overwrite earlier ones):

  1. NIH **Clinical Genomic Database (CGD)** — the authoritative, freely
     redistributable gene->inheritance table. Downloaded from research.nih.gov
     or read from a local file via ``--cgd-file`` (``.txt`` or ``.txt.gz``).
  2. A bundled curated seed (:data:`CURATED_INHERITANCE`) of well-established
     disease genes. This fills gaps and — importantly — lets the map be built
     when CGD is unavailable (e.g. offline environments).

Optionally, ``--clinvar-sqlite`` restricts/annotates output to the genes that
actually appear in a ClinVar SQLite (``variants.gene_symbol``), which is the
universe of genes the classifier will ever score.

Output format (consumed by ``local_db.inheritance_db.load_inheritance_map``)::

    gene<TAB>inheritance
    MVK<TAB>AR
    HBB<TAB>AR
    RHAG<TAB>AD/AR
    G6PD<TAB>XL

Inheritance codes follow CGD conventions: AD, AR, AD/AR, XL, XLR, XLD, MT
(mitochondrial), YL. ``inheritance_db.is_recessive`` grants the relaxed
(recessive) frequency threshold only to codes that are purely ``AR``/``XL``
(no AD component); combined codes such as ``AD/AR`` are treated as dominant.

Run standalone::

    python -m acmg_classifier.setup.build_inheritance_map \
        --clinvar-sqlite data/GRCh38/clinvar/clinvar_ps1_pm5_GRCh38.sqlite

    # offline (curated seed only):
    python -m acmg_classifier.setup.build_inheritance_map --no-download
"""
from __future__ import annotations

import argparse
import csv
import gzip
import io
import re
import sqlite3
import urllib.request
from pathlib import Path

import structlog

log = structlog.get_logger()

CGD_URL = "https://research.nih.gov/cgd/download/txt/CGD.txt.gz"

# CGD GENE/INHERITANCE column header candidates (CGD has used both casings).
_CGD_GENE_COLS = ("#GENE", "GENE", "Gene")
_CGD_INH_COLS = ("INHERITANCE", "Inheritance", "MOI")

# Inheritance tokens -> short codes. Covers both CGD short codes ("AD", "XL")
# and verbose phrases ("Autosomal dominant"). Non-MOI CGD tokens (BG = blood
# group, Digenic, Multigenic, PAR, Triallelic, Conditional) carry no
# mode-of-inheritance signal for AF thresholds and are simply dropped.
_TOKEN_TO_CODE = {
    "ad": "AD", "autosomal dominant": "AD",
    "ar": "AR", "autosomal recessive": "AR",
    "xl": "XL", "x-linked": "XL",
    "xlr": "XLR", "x-linked recessive": "XLR",
    "xld": "XLD", "x-linked dominant": "XLD",
    "yl": "YL", "y-linked": "YL",
    "mt": "MT", "mitochondrial": "MT", "maternal": "MT",
}
# Canonical output ordering: AD, AR, then X/Y/MT.
_CODE_ORDER = {"AD": 0, "AR": 1, "XL": 2, "XLR": 2, "XLD": 2, "YL": 3, "MT": 4}


# ---------------------------------------------------------------------------
# CGD acquisition + parsing
# ---------------------------------------------------------------------------

def download_cgd(url: str = CGD_URL, timeout: int = 30) -> bytes | None:
    """Download the gzipped CGD table. Returns raw (still gzipped) bytes or None."""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = resp.read()
        log.info("cgd_downloaded", bytes=len(data), url=url)
        return data
    except Exception as exc:  # noqa: BLE001 — network/IO failures are expected offline
        log.warning("cgd_download_failed", error=str(exc), url=url)
        return None


def _read_cgd_bytes(raw: bytes) -> str:
    """Decode CGD bytes, transparently gunzipping if needed."""
    if raw[:2] == b"\x1f\x8b":  # gzip magic
        raw = gzip.decompress(raw)
    return raw.decode("utf-8", "replace")


def _workbook_to_tsv_text(path: Path) -> str | None:
    """Convert the first sheet of an .xls/.xlsx workbook to tab-delimited text.

    .xls (legacy BIFF) needs ``xlrd``; .xlsx needs ``openpyxl``. Both are optional
    — imported lazily so the module works without them when CGD is supplied as
    plain text. Cell contents are flattened (tabs/newlines -> space) so the
    resulting rows stay cleanly tab-delimited for :func:`parse_cgd`.
    """
    suffix = path.suffix.lower()

    def _clean(v: object) -> str:
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        return str("" if v is None else v).replace("\t", " ").replace(
            "\r", " ").replace("\n", " ")

    try:
        if suffix == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            lines = ["\t".join(_clean(c) for c in row)
                     for row in ws.iter_rows(values_only=True)]
        else:  # .xls
            import xlrd
            wb = xlrd.open_workbook(path)
            sh = wb.sheet_by_index(0)
            lines = ["\t".join(_clean(sh.cell_value(r, c)) for c in range(sh.ncols))
                     for r in range(sh.nrows)]
    except ImportError as exc:
        log.error("xls_reader_missing", suffix=suffix, error=str(exc),
                  hint="pip install xlrd (.xls) / openpyxl (.xlsx)")
        return None
    except Exception as exc:  # noqa: BLE001
        log.error("xls_read_failed", path=str(path), error=str(exc))
        return None
    log.info("workbook_loaded", path=str(path), rows=len(lines))
    return "\n".join(lines)


def _load_cgd_text(cgd_file: Path | None, download: bool, url: str) -> str | None:
    if cgd_file is not None:
        if not cgd_file.exists():
            log.error("cgd_file_missing", path=str(cgd_file))
            return None
        if cgd_file.suffix.lower() in (".xls", ".xlsx"):
            return _workbook_to_tsv_text(cgd_file)
        return _read_cgd_bytes(cgd_file.read_bytes())
    if download:
        raw = download_cgd(url)
        return _read_cgd_bytes(raw) if raw is not None else None
    return None


def parse_cgd(text: str) -> dict[str, str]:
    """Parse CGD tab-delimited text into {gene_symbol: normalized_inheritance}."""
    reader = csv.DictReader(io.StringIO(text), delimiter="\t")
    fields = reader.fieldnames or []
    gene_col = next((c for c in _CGD_GENE_COLS if c in fields), None)
    inh_col = next((c for c in _CGD_INH_COLS if c in fields), None)
    if gene_col is None or inh_col is None:
        log.error("cgd_bad_header", fields=fields)
        return {}

    result: dict[str, str] = {}
    for row in reader:
        gene = (row.get(gene_col) or "").strip().lstrip("#").strip()
        code = normalize_inheritance(row.get(inh_col))
        if gene and code:
            result[gene] = code
    log.info("cgd_parsed", genes=len(result))
    return result


def normalize_inheritance(raw: str | None) -> str | None:
    """Normalize a raw/CGD inheritance string to short codes (e.g. ``AD/AR``).

    Handles CGD's real value space: combined codes (``AD/AR/Digenic``), verbose
    phrases (``Autosomal recessive``), parentheticals (``AD (with imprinting)``),
    casing (``Ad``), and non-MOI tokens (``BG``, ``Digenic``, ``PAR``) which are
    dropped. Returns None when no mode-of-inheritance token is present.
    """
    if not raw:
        return None
    # Drop parenthetical qualifiers, then lowercase for token matching.
    s = re.sub(r"\([^)]*\)", " ", str(raw)).strip().lower()
    if not s:
        return None

    codes: list[str] = []
    for tok in re.split(r"[\/,;]+|\s+and\s+", s):
        code = _TOKEN_TO_CODE.get(tok.strip())
        if code and code not in codes:
            codes.append(code)
    # Catch multi-word phrases that survive as trailing free text.
    for phrase, code in _TOKEN_TO_CODE.items():
        if " " in phrase and phrase in s and code not in codes:
            codes.append(code)
    if not codes:
        return None
    # A specific X-linked code (XLR/XLD) supersedes the generic XL.
    if "XLR" in codes or "XLD" in codes:
        codes = [c for c in codes if c != "XL"]
    codes = sorted(dict.fromkeys(codes), key=lambda c: _CODE_ORDER.get(c, 9))
    return "/".join(codes)


# ---------------------------------------------------------------------------
# ClinVar gene universe
# ---------------------------------------------------------------------------

def clinvar_genes(sqlite_path: Path) -> set[str]:
    """Return the set of distinct gene symbols present in a ClinVar SQLite."""
    if not sqlite_path.exists():
        log.warning("clinvar_sqlite_missing", path=str(sqlite_path))
        return set()
    try:
        con = sqlite3.connect(str(sqlite_path))
        rows = con.execute(
            "SELECT DISTINCT gene_symbol FROM variants WHERE gene_symbol IS NOT NULL"
        ).fetchall()
        con.close()
    except Exception as exc:  # noqa: BLE001
        log.error("clinvar_sqlite_error", error=str(exc))
        return set()
    genes = {(g[0] or "").strip() for g in rows if g[0]}
    log.info("clinvar_genes_loaded", genes=len(genes))
    return genes


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def build_inheritance_map(
    output: Path,
    cgd_file: Path | None = None,
    cgd_url: str = CGD_URL,
    download: bool = True,
    clinvar_sqlite: Path | None = None,
    restrict_to_clinvar: bool = False,
) -> dict[str, str]:
    """Build the gene->inheritance map and write it to ``output`` as TSV.

    The curated seed is the base layer; CGD entries override it where they
    overlap (CGD is the authoritative source). Returns the written map.
    """
    mapping: dict[str, str] = dict(CURATED_INHERITANCE)
    seed_n = len(mapping)

    text = _load_cgd_text(cgd_file, download, cgd_url)
    cgd_n = 0
    if text:
        cgd = parse_cgd(text)
        cgd_n = len(cgd)
        mapping.update(cgd)  # CGD wins on conflicts
    else:
        log.warning("cgd_unavailable_using_seed_only")

    cv_genes: set[str] = set()
    if clinvar_sqlite is not None:
        cv_genes = clinvar_genes(clinvar_sqlite)
        if cv_genes and restrict_to_clinvar:
            before = len(mapping)
            mapping = {g: c for g, c in mapping.items() if g in cv_genes}
            log.info("restricted_to_clinvar", before=before, after=len(mapping))

    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh, delimiter="\t", lineterminator="\n")
        writer.writerow(["gene", "inheritance"])
        for gene in sorted(mapping):
            writer.writerow([gene, mapping[gene]])

    covered = len(cv_genes & mapping.keys()) if cv_genes else 0
    log.info(
        "inheritance_map_written",
        path=str(output),
        total=len(mapping),
        seed=seed_n,
        cgd=cgd_n,
        clinvar_genes=len(cv_genes),
        clinvar_covered=covered,
    )
    return mapping


def main(argv: list[str] | None = None) -> int:
    structlog.configure(
        processors=[
            structlog.processors.TimeStamper(fmt="iso"),
            structlog.dev.ConsoleRenderer(),
        ]
    )
    p = argparse.ArgumentParser(description="Build gene_inheritance.tsv (gene -> MOI).")
    p.add_argument("--data-dir", type=Path, default=Path("./data"))
    p.add_argument("--output", type=Path, default=None,
                   help="Output TSV (default: resources/shared/gene_inheritance.tsv)")
    p.add_argument("--cgd-file", type=Path, default=None,
                   help="Local CGD.txt or CGD.txt.gz (skips download)")
    p.add_argument("--cgd-url", default=CGD_URL)
    p.add_argument("--no-download", action="store_true",
                   help="Do not fetch CGD; build from curated seed (+ --cgd-file)")
    p.add_argument("--clinvar-sqlite", type=Path, default=None,
                   help="ClinVar SQLite for coverage reporting / --restrict-to-clinvar")
    p.add_argument("--restrict-to-clinvar", action="store_true",
                   help="Emit only genes present in the ClinVar SQLite")
    args = p.parse_args(argv)

    # Curated table → committed resources/ tree (not the git-ignored data_dir).
    output = args.output or (Path("./resources") / "shared" / "gene_inheritance.tsv")
    build_inheritance_map(
        output=output,
        cgd_file=args.cgd_file,
        cgd_url=args.cgd_url,
        download=not args.no_download,
        clinvar_sqlite=args.clinvar_sqlite,
        restrict_to_clinvar=args.restrict_to_clinvar,
    )
    return 0


# ---------------------------------------------------------------------------
# Curated seed — well-established disease genes (OMIM / CGD / GeneReviews).
# Codes: AD, AR, AD/AR, XL, XLR, XLD, MT, YL. Kept intentionally conservative;
# CGD overrides these whenever it is available.
# ---------------------------------------------------------------------------
CURATED_INHERITANCE: dict[str, str] = {
    # --- Hereditary cancer ---
    "BRCA1": "AD", "BRCA2": "AD", "PALB2": "AD", "CHEK2": "AD", "ATM": "AD/AR",
    "TP53": "AD", "PTEN": "AD", "STK11": "AD", "CDH1": "AD", "NF1": "AD",
    "NF2": "AD", "RB1": "AD", "WT1": "AD", "VHL": "AD", "RET": "AD", "MEN1": "AD",
    "APC": "AD", "MUTYH": "AR", "MLH1": "AD", "MSH2": "AD", "MSH6": "AD",
    "PMS2": "AD", "EPCAM": "AD", "BMPR1A": "AD", "SMAD4": "AD", "GREM1": "AD",
    "POLE": "AD", "POLD1": "AD", "NTHL1": "AR", "AXIN2": "AD", "PTCH1": "AD",
    "SUFU": "AD", "DICER1": "AD", "HOXB13": "AD", "CDKN2A": "AD", "CDK4": "AD",
    "BAP1": "AD", "FLCN": "AD", "MET": "AD", "FH": "AD/AR", "SDHA": "AD",
    "SDHB": "AD", "SDHC": "AD", "SDHD": "AD", "SDHAF2": "AD", "MAX": "AD",
    "TMEM127": "AD", "TSC1": "AD", "TSC2": "AD", "RAD51C": "AD/AR",
    "RAD51D": "AD", "BRIP1": "AD/AR", "BARD1": "AD", "NBN": "AR",
    "GATA2": "AD", "RUNX1": "AD", "CEBPA": "AD", "DDX41": "AD", "ANKRD26": "AD",
    "ETV6": "AD", "WRN": "AR", "BLM": "AR", "RECQL4": "AR", "FANCA": "AR",
    "FANCC": "AR", "FANCG": "AR", "FANCD2": "AR", "SMARCB1": "AD", "SMARCA4": "AD",
    # --- Cardiomyopathy / arrhythmia ---
    "MYH7": "AD", "MYBPC3": "AD", "TNNT2": "AD", "TNNI3": "AD", "TPM1": "AD",
    "MYL2": "AD", "MYL3": "AD", "ACTC1": "AD", "TNNC1": "AD", "PLN": "AD",
    "CSRP3": "AD", "LMNA": "AD", "FLNC": "AD", "BAG3": "AD", "DES": "AD",
    "TTN": "AD", "RBM20": "AD", "SCN5A": "AD", "KCNQ1": "AD/AR", "KCNH2": "AD",
    "KCNE1": "AD/AR", "KCNE2": "AD", "RYR2": "AD", "CASQ2": "AD/AR",
    "CALM1": "AD", "CALM2": "AD", "CALM3": "AD", "TRDN": "AR", "ANK2": "AD",
    "PKP2": "AD", "DSP": "AD/AR", "DSG2": "AD", "DSC2": "AD/AR", "JUP": "AD/AR",
    "TMEM43": "AD", "SCN1B": "AD",
    # --- Aortopathy / connective tissue ---
    "FBN1": "AD", "TGFBR1": "AD", "TGFBR2": "AD", "SMAD3": "AD", "TGFB2": "AD",
    "TGFB3": "AD", "COL3A1": "AD", "ACTA2": "AD", "MYH11": "AD", "MYLK": "AD",
    "PRKG1": "AD", "LOX": "AD", "COL1A1": "AD", "COL1A2": "AD", "COL2A1": "AD",
    "COL11A1": "AD", "COL5A1": "AD", "COL5A2": "AD", "CRTAP": "AR", "P3H1": "AR",
    # --- Familial hypercholesterolemia / lipid ---
    "LDLR": "AD", "APOB": "AD", "PCSK9": "AD", "LDLRAP1": "AR",
    # --- Inborn errors of metabolism ---
    "PAH": "AR", "GAA": "AR", "HEXA": "AR", "HEXB": "AR", "GBA": "AR",
    "SMPD1": "AR", "ASAH1": "AR", "GALC": "AR", "ARSA": "AR", "GLA": "XL",
    "GBA1": "AR",
    "ABCD1": "XLR", "IDUA": "AR", "IDS": "XLR", "SGSH": "AR", "NAGLU": "AR",
    "GUSB": "AR", "GLB1": "AR", "NEU1": "AR", "CTNS": "AR", "ATP7B": "AR",
    "HFE": "AR", "SERPINA1": "AR", "MVK": "AR",
    "OTC": "XLR", "ASS1": "AR", "ASL": "AR", "CPS1": "AR", "ARG1": "AR",
    "NAGS": "AR", "SLC25A13": "AR", "MMUT": "AR", "MMAA": "AR", "MMAB": "AR",
    "PCCA": "AR", "PCCB": "AR", "MCCC1": "AR", "MCCC2": "AR", "IVD": "AR",
    "ACADM": "AR", "ACADVL": "AR", "ACADS": "AR", "HADHA": "AR", "HADHB": "AR",
    "CPT1A": "AR", "CPT2": "AR", "SLC22A5": "AR", "ETFA": "AR", "ETFB": "AR",
    "ETFDH": "AR", "GALT": "AR", "GALK1": "AR", "GBE1": "AR", "PYGL": "AR",
    "PYGM": "AR", "AGL": "AR", "SLC37A4": "AR", "G6PC1": "AR", "FBP1": "AR",
    "ALDOB": "AR", "GYS2": "AR", "G6PD": "XL",
    # --- Hemoglobinopathy / hematology / coagulation ---
    "HBB": "AR", "HBA1": "AR", "HBA2": "AR", "RHAG": "AD/AR", "SPTB": "AD",
    "ANK1": "AD", "F8": "XLR", "F9": "XLR", "VWF": "AD/AR", "F2": "AD",
    "F5": "AD/AR", "SERPINC1": "AD", "PROC": "AD/AR", "PROS1": "AD",
    # --- Cystic fibrosis / pulmonary ---
    "CFTR": "AR",
    # --- Renal ---
    "PKD1": "AD", "PKD2": "AD", "PKHD1": "AR", "NPHS1": "AR", "NPHS2": "AR",
    "UMOD": "AD", "COL4A3": "AD/AR", "COL4A4": "AD/AR", "COL4A5": "XL",
    # --- Neuromuscular ---
    "DMD": "XLR", "SMN1": "AR", "FMR1": "XL", "DMPK": "AD", "CACNA1S": "AD/AR",
    "RYR1": "AD/AR", "PMP22": "AD", "MPZ": "AD", "GJB1": "XL", "MFN2": "AD",
    "SPAST": "AD", "SPG7": "AR", "SOD1": "AD", "C9orf72": "AD", "TARDBP": "AD",
    "FUS": "AD",
    # --- Neurodevelopmental / epilepsy ---
    "MECP2": "XLD", "CDKL5": "XL", "FOXG1": "AD", "UBE3A": "AD", "SHANK3": "AD",
    "SCN1A": "AD", "SCN2A": "AD", "SCN8A": "AD", "KCNQ2": "AD", "STXBP1": "AD",
    "GABRA1": "AD", "GRIN2A": "AD", "CACNA1A": "AD", "KCNT1": "AD", "GNAO1": "AD",
    "TCF4": "AD", "ARID1B": "AD", "ADNP": "AD", "SYNGAP1": "AD",
    # --- Movement / neurodegeneration ---
    "HTT": "AD", "ATXN1": "AD", "ATXN2": "AD", "ATXN3": "AD", "SNCA": "AD",
    "LRRK2": "AD", "PRKN": "AR", "PINK1": "AR", "PARK7": "AR",
    # --- Ophthalmology ---
    "RHO": "AD", "RPGR": "XL", "RP1": "AD/AR", "USH2A": "AR", "ABCA4": "AR",
    "CRB1": "AR", "RPE65": "AR", "CEP290": "AR", "BBS1": "AR", "BBS2": "AR",
    "MYO7A": "AR",
    # --- Skeletal / craniofacial ---
    "FGFR1": "AD", "FGFR2": "AD", "FGFR3": "AD", "RUNX2": "AD", "SOX9": "AD",
    # --- Immunodeficiency ---
    "BTK": "XLR", "IL2RG": "XLR", "WAS": "XLR", "ADA": "AR", "CYBB": "XLR",
    # --- Amyloidosis / other ---
    "TTR": "AD",
}


if __name__ == "__main__":
    raise SystemExit(main())
